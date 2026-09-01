# -*- coding: utf-8 -*-
"""统一表格数据源：读取 CSV / XLSX / XLS / XLSB，提供 names() / header(sheet) / all_rows(sheet)。
所有数据源均采用懒加载+缓存；本模块不依赖其他功能模块，可独立修改。"""
import os
import csv
import io

VALID_EXTS = {'.csv', '.xlsx', '.xlsm', '.xls', '.xlsb'}

CSV_ENC = ['utf-8-sig', 'gbk', 'utf-8']


class _BaseSource(object):
    def names(self):
        raise NotImplementedError

    def header(self, sheet):
        raise NotImplementedError

    def all_rows(self, sheet):
        raise NotImplementedError


class _OpenPyXlSource(_BaseSource):
    def __init__(self, path):
        import openpyxl
        self._path = path
        self._name = None

    def names(self):
        if self._name is None:
            import openpyxl
            wb = openpyxl.load_workbook(self._path, read_only=True, data_only=True)
            self._name = list(wb.sheetnames)
        return self._name

    def header(self, sheet):
        import openpyxl
        wb = openpyxl.load_workbook(self._path, read_only=True, data_only=True)
        for row in wb[sheet].iter_rows(values_only=True):
            if any(c is not None for c in row):
                return [str(c) if c is not None else '' for c in row]
        return []

    def all_rows(self, sheet):
        import openpyxl
        wb = openpyxl.load_workbook(self._path, read_only=True, data_only=True)
        ws = wb[sheet]
        header = None
        rows = []
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = [str(c) if c is not None else '' for c in row]
                continue
            if not any(c is not None for c in row):
                continue
            rows.append({header[i]: row[i] for i in range(len(header))})
        return rows


class _PandasSource(_BaseSource):
    """用于 .xls（依赖 pandas+xlrd）。"""

    def __init__(self, path):
        import pandas as pd
        self._ef = pd.ExcelFile(path)

    def names(self):
        return list(self._ef.sheet_names)

    def header(self, sheet):
        df = self._ef.parse(sheet, nrows=1)
        return [str(c) if c is not None else '' for c in df.columns]

    def all_rows(self, sheet):
        df = self._ef.parse(sheet)
        header = [str(c) if c is not None else '' for c in df.columns]
        rows = []
        for r in df.to_dict('records'):
            rows.append(dict(zip(header, list(r.values()))))
        return rows


class _XlsbSource(_BaseSource):
    """用于 .xlsb（依赖 pyxlsb）。"""

    def __init__(self, path):
        try:
            from pyxlsb import open_workbook
        except ImportError:
            raise ImportError('读取 XLSB 需要安装 pyxlsb，请执行: pip install pyxlsb')
        self._path = path
        self._wb = open_workbook(path)
        self._name = list(self._wb.sheets)

    def names(self):
        return self._name

    def header(self, sheet):
        ws = self._wb.get_sheet(sheet)
        for row in ws.rows():
            vals = [c.v for c in row]
            if any(v is not None for v in vals):
                return [str(v) if v is not None else '' for v in vals]
        return []

    def all_rows(self, sheet):
        ws = self._wb.get_sheet(sheet)
        header = None
        rows = []
        for row in ws.rows():
            vals = [c.v for c in row]
            if header is None:
                header = [str(v) if v is not None else '' for v in vals]
                continue
            if all(v is None for v in vals):
                continue
            rows.append({header[i]: vals[i] for i in range(len(header))})
        return rows


class _CsvSource(_BaseSource):
    def __init__(self, path):
        self._path = path
        self._name = ['CSV']
        self._headers = None
        self._rows = None
        self._sheet_loaded = None

    def names(self):
        return self._name

    def _read_text(self, encoding):
        with open(self._path, 'r', encoding=encoding, newline='') as f:
            return f.read()

    def _load(self):
        """读取全部，自动尝试编码与分隔符。"""
        text = None
        for enc in CSV_ENC:
            try:
                text = self._read_text(enc)
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
        if text is None:
            raise IOError('无法识别 CSV 编码（已尝试 utf-8/gbk）')
        # 分隔符检测
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
            delim = dialect.delimiter
        except Exception:
            delim = ','
        reader = csv.reader(io.StringIO(text), delimiter=delim)
        raw = [row for row in reader if row and any(c.strip() for c in row)]
        header = [c.strip() for c in raw[0]] if raw else []
        self._headers = header
        rows = []
        for r in raw[1:]:
            rec = {}
            for i, h in enumerate(header):
                rec[h] = r[i].strip() if i < len(r) else ''
            rows.append(rec)
        self._rows = rows

    def header(self, sheet):
        if self._headers is None:
            self._load()
        return self._headers

    def all_rows(self, sheet):
        if self._rows is None:
            self._load()
        return self._rows


def open_source(path):
    """返回 (source, error)。"""
    ext = os.path.splitext(path)[1].lower()
    if ext not in VALID_EXTS:
        return None, '不支持的文件格式: %s' % ext
    try:
        if ext == '.csv':
            src = _CsvSource(path)
        elif ext in ('.xlsx', '.xlsm'):
            src = _OpenPyXlSource(path)
        elif ext == '.xls':
            src = _PandasSource(path)
        else:  # .xlsb
            src = _XlsbSource(path)
        # 触发一次 names() 以尽早暴露依赖缺失等问题
        src.names()
        return src, None
    except Exception as exc:
        return None, str(exc)